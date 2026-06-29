from collections.abc import Iterator
from io import BytesIO
from zipfile import ZipFile

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.core.config import get_settings
from app.core.database import get_engine
from app.main import create_application
from app.seed import seed_database


@pytest.fixture
def client(tmp_path, monkeypatch) -> Iterator[TestClient]:
    database_path = tmp_path / 'imports-test.db'
    monkeypatch.setenv('DATABASE_URL', f'sqlite:///{database_path}')
    monkeypatch.setenv('AUTH_BOOTSTRAP_EMAIL', 'admin@example.com')
    monkeypatch.setenv('AUTH_BOOTSTRAP_PASSWORD', 'super-secret-password')
    monkeypatch.setenv('AUTH_BOOTSTRAP_FULL_NAME', 'Test Admin')
    monkeypatch.setenv('AUTH_COOKIE_SECURE', 'false')

    get_settings.cache_clear()
    get_engine.cache_clear()

    try:
        with TestClient(create_application()) as test_client:
            yield test_client
    finally:
        get_engine.cache_clear()
        get_settings.cache_clear()


def test_import_pdx_workbook_creates_and_updates_records(client: TestClient):
    login_response = client.post(
        '/api/auth/login',
        json={'email': 'admin@example.com', 'password': 'super-secret-password'},
    )
    assert login_response.status_code == 200

    first_workbook = _build_workbook(age=63, diagnosis='Adenocarcinoma de pulmón')

    first_response = client.post(
        '/api/imports/pdx-workbook',
        files={
            'file': (
                'PDXs.xlsx',
                first_workbook.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        },
    )

    assert first_response.status_code == 200
    assert first_response.json() == {
        'filename': 'PDXs.xlsx',
        'sheets_processed': 1,
        'rows_imported': 1,
        'rows_skipped': 0,
        'rows_failed': 0,
        'patients': {'created': 1, 'updated': 0},
        'tumors': {'created': 1, 'updated': 0},
        'biomodels': {'created': 1, 'updated': 0},
        'passages': {'created': 2, 'updated': 0},
        'pdx_trials': {'created': 2, 'updated': 0},
        'images': {'created': 2, 'updated': 0},
        'errors': [],
    }

    patient_response = client.get('/api/patients/88888')
    assert patient_response.status_code == 200
    assert patient_response.json()['sex'] == 'F'
    assert patient_response.json()['age'] == 63

    tumor_response = client.get('/api/tumors/BIOBPRQ1246')
    assert tumor_response.status_code == 200
    assert tumor_response.json()['tube_code'] == '44000008'
    assert tumor_response.json()['ap_diagnosis'] == 'Adenocarcinoma de pulmón'
    assert tumor_response.json()['intervention_date'] == '2021-02-09'

    biomodel_response = client.get('/api/biomodels/LUNG 090221')
    assert biomodel_response.status_code == 200
    assert biomodel_response.json()['type'] == 'PDX'
    assert biomodel_response.json()['tumor_biobank_code'] == 'BIOBPRQ1246'

    passages_response = client.get('/api/passages')
    assert passages_response.status_code == 200
    assert {passage['id'] for passage in passages_response.json()} == {
        'LUNG 090221-P1',
        'LUNG 090221-P2',
    }

    pdx_trials_response = client.get('/api/pdx-trials')
    assert pdx_trials_response.status_code == 200
    assert {trial['id'] for trial in pdx_trials_response.json()} == {
        'LUNG 090221-P1',
        'LUNG 090221-P2',
    }

    images_response = client.get('/api/images')
    assert images_response.status_code == 200
    images = sorted(images_response.json(), key=lambda image: image['passage_id'])
    assert len(images) == 2
    assert images[0]['passage_id'] == 'LUNG 090221-P1'
    assert images[0]['type'] == 'H&E'
    assert images[0]['scanner_magnification'] == 40
    assert images[0]['image_date'] == '2021-10-14'
    assert images[1]['passage_id'] == 'LUNG 090221-P2'
    assert images[1]['type'] == 'H&E'
    assert images[1]['scanner_magnification'] == 40
    assert images[1]['image_date'] == '2022-04-21'

    second_workbook = _build_workbook(age=64, diagnosis='Diagnóstico actualizado')
    second_response = client.post(
        '/api/imports/pdx-workbook',
        files={
            'file': (
                'PDXs.xlsx',
                second_workbook.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        },
    )

    assert second_response.status_code == 200
    assert second_response.json()['patients'] == {'created': 0, 'updated': 1}
    assert second_response.json()['tumors'] == {'created': 0, 'updated': 1}
    assert second_response.json()['biomodels'] == {'created': 0, 'updated': 1}
    assert second_response.json()['passages'] == {'created': 0, 'updated': 2}
    assert second_response.json()['pdx_trials'] == {'created': 0, 'updated': 2}
    assert second_response.json()['images'] == {'created': 0, 'updated': 2}

    updated_patient = client.get('/api/patients/88888')
    assert updated_patient.status_code == 200
    assert updated_patient.json()['age'] == 64

    updated_tumor = client.get('/api/tumors/BIOBPRQ1246')
    assert updated_tumor.status_code == 200
    assert updated_tumor.json()['ap_diagnosis'] == 'Diagnóstico actualizado'


def test_import_pdx_workbook_supports_legacy_sheet_layout(client: TestClient):
    login_response = client.post(
        '/api/auth/login',
        json={'email': 'admin@example.com', 'password': 'super-secret-password'},
    )
    assert login_response.status_code == 200

    workbook = _build_legacy_workbook()
    response = client.post(
        '/api/imports/pdx-workbook',
        files={
            'file': (
                'legacy.xlsx',
                workbook.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        },
    )

    assert response.status_code == 200
    assert response.json()['rows_imported'] == 1
    assert response.json()['passages'] == {'created': 2, 'updated': 0}
    assert response.json()['images'] == {'created': 1, 'updated': 0}

    biomodel_response = client.get('/api/biomodels/SAR011124')
    assert biomodel_response.status_code == 200
    assert biomodel_response.json()['tumor_biobank_code'] == 'BIOBSAR0007'

    passages_response = client.get('/api/passages')
    assert passages_response.status_code == 200
    passages = sorted(passages_response.json(), key=lambda passage: passage['id'])
    assert passages[0]['id'] == 'SAR011124-P1'
    assert passages[0]['success'] is False
    assert passages[1]['id'] == 'SAR011124-P2'
    assert passages[1]['success'] is True

    images_response = client.get('/api/images')
    assert images_response.status_code == 200
    images = images_response.json()
    assert len(images) == 1
    assert images[0]['passage_id'] == 'SAR011124-P2'
    assert images[0]['type'] == 'H&E'
    assert images[0]['scanner_magnification'] == 40
    assert images[0]['image_date'] == '2024-12-05'


def test_download_dataset_template_workbook(client: TestClient):
    login_response = client.post(
        '/api/auth/login',
        json={'email': 'admin@example.com', 'password': 'super-secret-password'},
    )
    assert login_response.status_code == 200

    response = client.get('/api/imports/dataset-template.xlsx')

    assert response.status_code == 200
    assert response.headers['content-type'].startswith(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    assert 'techconnect-dataset-template.xlsx' in response.headers['content-disposition']

    workbook = load_workbook(BytesIO(response.content), read_only=True)
    try:
        assert 'README' not in workbook.sheetnames
        assert workbook.sheetnames[0] == 'patient'
        assert 'patient' in workbook.sheetnames
        assert 'tumor' in workbook.sheetnames
        assert 'biomodel' in workbook.sheetnames
        assert 'mouse' in workbook.sheetnames
        assert 'pdx_passage' in workbook.sheetnames
        assert 'pdo_passage' in workbook.sheetnames
        assert 'lc_passage' in workbook.sheetnames
        assert 'pdx_trial' not in workbook.sheetnames
        assert 'pdo_trial' not in workbook.sheetnames
        assert 'lc_trial' not in workbook.sheetnames
        assert 'implant' not in workbook.sheetnames
        assert 'measure' not in workbook.sheetnames
        assert 'implant_measure' not in workbook.sheetnames

        patient_headers = next(workbook['patient'].iter_rows(min_row=1, max_row=1, values_only=True))
        patient_notes = next(workbook['patient'].iter_rows(min_row=2, max_row=2, values_only=True))
        tumor_headers = next(workbook['tumor'].iter_rows(min_row=1, max_row=1, values_only=True))
        mouse_headers = next(workbook['mouse'].iter_rows(min_row=1, max_row=1, values_only=True))

        assert patient_headers == ('nhc', 'sex', 'age')
        assert patient_notes[0] == 'primary key | required | type:string'
        assert tumor_headers[-1] == 'patient_nhc'
        assert mouse_headers[-8:] == (
            'implant_1_id',
            'implant_1_location',
            'implant_1_type',
            'implant_1_date',
            'implant_2_id',
            'implant_2_location',
            'implant_2_type',
            'implant_2_date',
        )
    finally:
        workbook.close()


def test_download_dataset_template_csv_zip(client: TestClient):
    login_response = client.post(
        '/api/auth/login',
        json={'email': 'admin@example.com', 'password': 'super-secret-password'},
    )
    assert login_response.status_code == 200

    response = client.get('/api/imports/dataset-template.zip')

    assert response.status_code == 200
    assert response.headers['content-type'].startswith('application/zip')
    assert 'techconnect-dataset-template.zip' in response.headers['content-disposition']

    with ZipFile(BytesIO(response.content)) as archive:
        names = set(archive.namelist())
        assert 'README.txt' in names
        assert 'patient.csv' in names
        assert 'tumor.csv' in names
        assert 'biomodel.csv' in names
        assert 'mouse.csv' in names
        assert 'pdx_passage.csv' in names
        assert 'pdo_passage.csv' in names
        assert 'lc_passage.csv' in names
        assert 'pdx_trial.csv' not in names
        assert 'pdo_trial.csv' not in names
        assert 'lc_trial.csv' not in names
        assert 'implant.csv' not in names
        assert 'measure.csv' not in names
        assert 'implant_measure.csv' not in names
        patient_csv = archive.read('patient.csv').decode('utf-8').strip()
        readme = archive.read('README.txt').decode('utf-8')

    assert patient_csv == 'nhc,sex,age'
    assert 'Dates should use YYYY-MM-DD.' in readme


def test_export_dataset_workbook_includes_existing_rows(client: TestClient):
    login_response = client.post(
        '/api/auth/login',
        json={'email': 'admin@example.com', 'password': 'super-secret-password'},
    )
    assert login_response.status_code == 200

    patient_response = client.post('/api/patients', json={'nhc': 'PAT-001', 'sex': 'F', 'age': 42})
    assert patient_response.status_code == 200
    tumor_response = client.post(
        '/api/tumors',
        json={
            'biobank_code': 'TUM-001',
            'patient_nhc': 'PAT-001',
            'organ': 'Lung',
            'classification': 'Adenocarcinoma',
        },
    )
    assert tumor_response.status_code == 200

    response = client.get('/api/imports/dataset.xlsx')

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    try:
        assert 'README' not in workbook.sheetnames
        assert workbook.sheetnames[0] == 'patient'
        assert 'measure' in workbook.sheetnames
        assert 'implant' not in workbook.sheetnames
        patient_rows = list(workbook['patient'].iter_rows(min_row=3, values_only=True))
        tumor_rows = list(workbook['tumor'].iter_rows(min_row=3, values_only=True))
    finally:
        workbook.close()

    assert ('PAT-001', 'F', 42) in patient_rows
    assert ('TUM-001', None, 'Adenocarcinoma', None, None, 'Lung', None, None, None, 'PAT-001') in tumor_rows


def test_import_dataset_workbook_creates_and_updates_related_rows(client: TestClient):
    login_response = client.post(
        '/api/auth/login',
        json={'email': 'admin@example.com', 'password': 'super-secret-password'},
    )
    assert login_response.status_code == 200

    template_response = client.get('/api/imports/dataset-template.xlsx')
    workbook = load_workbook(BytesIO(template_response.content))
    patient_sheet = workbook['patient']
    tumor_sheet = workbook['tumor']
    patient_sheet.append(['PAT-100', 'F', 35])
    tumor_sheet.append(['TUM-100', None, None, 'Adenocarcinoma', None, 'Lung', None, None, None, 'PAT-100'])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()

    first_response = client.post(
        '/api/imports/dataset-workbook',
        files={
            'file': (
                'dataset.xlsx',
                payload.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        },
    )

    assert first_response.status_code == 200
    assert first_response.json()['rows_imported'] == 2
    assert first_response.json()['table_counts']['patient'] == {'created': 1, 'updated': 0}
    assert first_response.json()['table_counts']['tumor'] == {'created': 1, 'updated': 0}

    patient_response = client.get('/api/patients/PAT-100')
    assert patient_response.status_code == 200
    assert patient_response.json()['age'] == 35

    updated_workbook = load_workbook(BytesIO(template_response.content))
    updated_patient_sheet = updated_workbook['patient']
    updated_tumor_sheet = updated_workbook['tumor']
    updated_patient_sheet.append(['PAT-100', 'F', 36])
    updated_tumor_sheet.append(['TUM-100', None, None, 'Updated diagnosis', None, 'Lung', None, None, None, 'PAT-100'])
    updated_payload = BytesIO()
    updated_workbook.save(updated_payload)
    updated_workbook.close()

    second_response = client.post(
        '/api/imports/dataset-workbook',
        files={
            'file': (
                'dataset.xlsx',
                updated_payload.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        },
    )

    assert second_response.status_code == 200
    assert second_response.json()['table_counts']['patient'] == {'created': 0, 'updated': 1}
    assert second_response.json()['table_counts']['tumor'] == {'created': 0, 'updated': 1}

    updated_patient_response = client.get('/api/patients/PAT-100')
    updated_tumor_response = client.get('/api/tumors/TUM-100')
    assert updated_patient_response.json()['age'] == 36
    assert updated_tumor_response.json()['ap_diagnosis'] == 'Updated diagnosis'


def test_import_dataset_workbook_coerces_numeric_excel_values_for_string_columns(client: TestClient):
    login_response = client.post(
        '/api/auth/login',
        json={'email': 'admin@example.com', 'password': 'super-secret-password'},
    )
    assert login_response.status_code == 200

    template_response = client.get('/api/imports/dataset-template.xlsx')
    workbook = load_workbook(BytesIO(template_response.content))
    patient_sheet = workbook['patient']
    tumor_sheet = workbook['tumor']
    patient_sheet.append(['PAT-300', 'F', 35])
    tumor_sheet.append(['TUM-300', 222222222, 'carcinoma', None, 3, 'pulmon', 2, 'T1N2M0', None, 'PAT-300'])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()

    response = client.post(
        '/api/imports/dataset-workbook',
        files={
            'file': (
                'dataset.xlsx',
                payload.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body['rows_failed'] == 0
    assert body['table_counts']['tumor'] == {'created': 1, 'updated': 0}

    tumor_response = client.get('/api/tumors/TUM-300')
    assert tumor_response.status_code == 200
    tumor = tumor_response.json()
    assert tumor['tube_code'] == '222222222'
    assert tumor['grade'] == '3'
    assert tumor['stage'] == '2'


def test_import_dataset_workbook_normalizes_passage_identifier_spaces(client: TestClient):
    login_response = client.post(
        '/api/auth/login',
        json={'email': 'admin@example.com', 'password': 'super-secret-password'},
    )
    assert login_response.status_code == 200

    template_response = client.get('/api/imports/dataset-template.xlsx')
    workbook = load_workbook(BytesIO(template_response.content))
    patient_sheet = workbook['patient']
    tumor_sheet = workbook['tumor']
    biomodel_sheet = workbook['biomodel']
    passage_sheet = workbook['passage']
    pdx_trial_sheet = workbook['pdx_passage']

    patient_sheet.append(['PAT-400', 'F', 35])
    tumor_sheet.append(['TUM-400', None, None, 'Adenocarcinoma', None, 'Lung', None, None, None, 'PAT-400'])
    biomodel_sheet.append(['LUNG260526', 'PDX', None, None, None, None, 'TUM-400', None])
    passage_sheet.append(['LUNG260526 PX2', None, 'YES', 'YES', None, 'NO', None, 'LUNG260526'])
    pdx_trial_sheet.append(['LUNG260526 PX2', None, None, None, None, None])

    payload = BytesIO()
    workbook.save(payload)
    workbook.close()

    response = client.post(
        '/api/imports/dataset-workbook',
        files={
            'file': (
                'dataset.xlsx',
                payload.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body['rows_failed'] == 0
    assert body['table_counts']['passage'] == {'created': 1, 'updated': 0}
    assert body['table_counts']['pdx_trial'] == {'created': 0, 'updated': 1}

    passage_response = client.get('/api/passages/LUNG260526-PX2')
    pdx_trial_response = client.get('/api/pdx-trials/LUNG260526-PX2')
    assert passage_response.status_code == 200
    assert pdx_trial_response.status_code == 200


def test_import_dataset_workbook_creates_mouse_implants_from_mouse_sheet(client: TestClient):
    login_response = client.post(
        '/api/auth/login',
        json={'email': 'admin@example.com', 'password': 'super-secret-password'},
    )
    assert login_response.status_code == 200

    template_response = client.get('/api/imports/dataset-template.xlsx')
    workbook = load_workbook(BytesIO(template_response.content))
    patient_sheet = workbook['patient']
    tumor_sheet = workbook['tumor']
    biomodel_sheet = workbook['biomodel']
    passage_sheet = workbook['passage']
    pdx_trial_sheet = workbook['pdx_passage']
    mouse_sheet = workbook['mouse']

    patient_sheet.append(['PAT-500', 'F', 35])
    tumor_sheet.append(['TUM-500', None, None, 'Adenocarcinoma', None, 'Lung', None, None, None, 'PAT-500'])
    biomodel_sheet.append(['LUNG500', 'PDX', None, None, None, None, 'TUM-500', None])
    passage_sheet.append(['LUNG500 PX2', None, 'YES', 'YES', None, 'NO', None, 'LUNG500'])
    pdx_trial_sheet.append(['LUNG500 PX2', None, None, None, None, None])
    mouse_sheet.append([
        None,
        None,
        None,
        'AF-1',
        'PROEX-1',
        'NSG',
        'female',
        None,
        'LUNG500 PX2',
        None,
        'izquierda',
        'subcutaneo',
        '2023-04-15',
        None,
        'derecha',
        'subcutaneo',
        '2023-04-20',
    ])

    payload = BytesIO()
    workbook.save(payload)
    workbook.close()

    response = client.post(
        '/api/imports/dataset-workbook',
        files={
            'file': (
                'dataset.xlsx',
                payload.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body['rows_failed'] == 0
    assert body['table_counts']['mouse'] == {'created': 1, 'updated': 0}
    assert body['table_counts']['implant'] == {'created': 2, 'updated': 0}

    mice_response = client.get('/api/mice')
    implants_response = client.get('/api/implants')
    assert mice_response.status_code == 200
    assert implants_response.status_code == 200

    mice = mice_response.json()
    implants = implants_response.json()
    assert len(mice) == 1
    assert mice[0]['pdx_trial_id'] == 'LUNG500-PX2'
    assert {implant['mouse_id'] for implant in implants} == {mice[0]['id']}
    assert {implant['implant_location'] for implant in implants} == {'izquierda', 'derecha'}
    assert {implant['implant_date'] for implant in implants} == {'2023-04-15', '2023-04-20'}


def test_exported_dataset_workbook_roundtrips_seed_data(client: TestClient):
    login_response = client.post(
        '/api/auth/login',
        json={'email': 'admin@example.com', 'password': 'super-secret-password'},
    )
    assert login_response.status_code == 200

    seed_stats = seed_database()
    assert seed_stats.created > 0

    export_response = client.get('/api/imports/dataset.xlsx')
    assert export_response.status_code == 200

    import_response = client.post(
        '/api/imports/dataset-workbook',
        files={
            'file': (
                'techconnect-dataset.xlsx',
                export_response.content,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        },
    )

    assert import_response.status_code == 200
    body = import_response.json()
    assert body['rows_failed'] == 0
    assert body['errors'] == []
    assert body['table_counts']['mouse']['updated'] > 0


def test_import_dataset_csv_zip_reports_partial_failures(client: TestClient):
    login_response = client.post(
        '/api/auth/login',
        json={'email': 'admin@example.com', 'password': 'super-secret-password'},
    )
    assert login_response.status_code == 200

    archive_buffer = BytesIO()
    with ZipFile(archive_buffer, mode='w') as archive:
        archive.writestr('patient.csv', 'nhc,sex,age\nPAT-200,F,40\n')
        archive.writestr('tumor.csv', 'biobank_code,tube_code,classification,ap_diagnosis,grade,organ,stage,tnm,intervention_date,patient_nhc\nTUM-200,,Adenocarcinoma,,,Lung,,,,PAT-200\nTUM-201,,Adenocarcinoma,,,Lung,,,,MISSING\n')

    response = client.post(
        '/api/imports/dataset-csv-zip',
        files={'file': ('dataset.zip', archive_buffer.getvalue(), 'application/zip')},
    )

    assert response.status_code == 200
    body = response.json()
    assert body['rows_imported'] == 2
    assert body['rows_failed'] == 1
    assert body['table_counts']['patient'] == {'created': 1, 'updated': 0}
    assert body['table_counts']['tumor'] == {'created': 1, 'updated': 0}
    assert body['errors'][0]['table'] == 'tumor'
    assert body['errors'][0]['primary_key'] == 'TUM-201'

    patient_response = client.get('/api/patients/PAT-200')
    tumor_response = client.get('/api/tumors/TUM-200')
    missing_tumor_response = client.get('/api/tumors/TUM-201')
    assert patient_response.status_code == 200
    assert tumor_response.status_code == 200
    assert missing_tumor_response.status_code == 404


def _build_workbook(*, age: int, diagnosis: str) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'LUNG'

    worksheet.append(
        [
            'TUMOR',
            None,
            None,
            None,
            None,
            'BIOBANCO CODE',
            'DATE SURGERY',
            'NHC',
            'CÓDIGO',
            'DIAGNÓSTICO AP BIOPSIA',
            'SEXO',
            'EDAD',
            None,
            None,
            None,
            None,
            'px0',
            'SAMPLES',
            'FFPE',
            'Slides',
            'Scanner',
            'BB SAMPLES',
            'px1',
            'SAMPLES',
            'FFPE',
            'Slides',
            'Scanner',
            'BB SAMPLES',
        ]
    )
    worksheet.append(
        [
            'BIOMODEL. ID BIOMODELOS',
            'NUEVO TUMOR. Grado',
            'NUEVO TUMOR. Órgano',
            'NUEVO TUMOR. Estadio',
            'NUEVO TUMOR. TNM',
            'NUEVO TUMOR. codigo de tubo',
            'NUEVO TUMOR. Fecha de operación.',
            'PACIENTE. NHC',
            'NUEVO TUMOR. Código de Biobanco',
            'NUEVO TUMOR. Diagnóstico AP',
            'PACIENTE. SEXO',
            'PACIENTE. EDAD',
        ]
    )
    worksheet.append(
        [
            'LUNG 090221',
            None,
            'Lung',
            'IIIA',
            'T2N2M0',
            '44000008',
            '09/02/2021',
            '88888',
            'BIOBPRQ1246',
            diagnosis,
            'M',
            age,
            None,
            None,
            None,
            None,
            '✅',
            'CV',
            '✅',
            'H&E',
            '40x',
            '14/10/2021',
            '✅',
            None,
            '✅',
            'H&E',
            '40x',
            '21/04/2022',
        ]
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    workbook.close()
    return output


def _build_legacy_workbook() -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'SAR'

    worksheet.append(
        [
            'TUMOR',
            'BIOBANCO CODE',
            'DATE SURGERY',
            'NHC',
            'CÓDIGO',
            'DIAGNÓSTICO AP BIOPSIA',
            'SEXO',
            'EDAD',
            'INCIDENCIAS',
            'px0',
            'SAMPLES',
            'FFPE',
            'Slides',
            'Scanner',
            'BB SAMPLES',
            'px1',
            'SAMPLES',
            'FFPE',
            'Slides',
            'Scanner',
            'BB SAMPLES',
        ]
    )
    worksheet.append(
        [
            'SAR011124',
            4400006,
            '01/11/2024',
            33333,
            'BIOBSAR0007',
            'Liposarcoma mixoide de bajo grado',
            'V',
            46,
            'implantado el 05/12/24',
            '❌',
            None,
            None,
            None,
            None,
            None,
            '✅',
            'CV',
            '✅',
            'H&E',
            '40x',
            '05/12/2024',
        ]
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    workbook.close()
    return output