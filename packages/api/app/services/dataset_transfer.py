"""Generic dataset transfer helpers for workbook and CSV templates."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any
from uuid import UUID
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi import HTTPException, status
from models import Implant, Measure, Mouse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pydantic import BaseModel, Field, ValidationError
from sqlalchemy.exc import SQLAlchemyError
from sqlmodel import Session, SQLModel, select

from app.services.crud import create_item, update_item
from app.services.entity_catalog import ENTITY_ROUTERS


@dataclass(frozen=True)
class DatasetColumnSpec:
    name: str
    required: bool
    is_primary_key: bool
    foreign_keys: tuple[str, ...]
    data_type: str
    format_hint: str | None = None


@dataclass(frozen=True)
class DatasetTableSpec:
    model: type[SQLModel]
    route_prefix: str
    tag: str
    table_name: str
    columns: tuple[DatasetColumnSpec, ...]


class EntityImportCounts(BaseModel):
    created: int = 0
    updated: int = 0


class DatasetImportError(BaseModel):
    table: str
    row_number: int
    primary_key: str | None = None
    message: str


class DatasetImportSummary(BaseModel):
    filename: str | None = None
    format: str
    tables_processed: int = 0
    rows_imported: int = 0
    rows_skipped: int = 0
    rows_failed: int = 0
    table_counts: dict[str, EntityImportCounts] = Field(default_factory=dict)
    errors: list[DatasetImportError] = Field(default_factory=list)


DATASET_SHEET_NAME_ALIASES = {
    "pdx_trial": "pdx_passage",
    "pdo_trial": "pdo_passage",
    "lc_trial": "lc_passage",
}


MOUSE_RELATED_COLUMNS: tuple[DatasetColumnSpec, ...] = (
    DatasetColumnSpec("implant_1_id", False, False, (), "uuid", "xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx"),
    DatasetColumnSpec("implant_1_location", False, False, (), "string"),
    DatasetColumnSpec("implant_1_type", False, False, (), "string"),
    DatasetColumnSpec("implant_1_date", False, False, (), "date", "YYYY-MM-DD"),
    DatasetColumnSpec("implant_2_id", False, False, (), "uuid", "xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx"),
    DatasetColumnSpec("implant_2_location", False, False, (), "string"),
    DatasetColumnSpec("implant_2_type", False, False, (), "string"),
    DatasetColumnSpec("implant_2_date", False, False, (), "date", "YYYY-MM-DD"),
)


def get_dataset_table_specs() -> tuple[DatasetTableSpec, ...]:
    """Return ordered dataset table specs for all supported domain entities."""
    excluded_tables = {"implant", "measure"}
    return tuple(
        DatasetTableSpec(
            model=model,
            route_prefix=route_prefix,
            tag=tag,
            table_name=model.__table__.name,
            columns=_get_model_columns(model),
        )
        for model, route_prefix, tag in ENTITY_ROUTERS
        if model.__table__.name not in excluded_tables
    )


def build_dataset_template_workbook() -> bytes:
    """Create an empty workbook template with one sheet per domain table."""
    workbook = _build_base_workbook()
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def build_dataset_export_workbook(session: Session) -> bytes:
    """Create a workbook export populated with current database contents."""
    workbook = _build_base_workbook()

    for table_spec in get_dataset_table_specs():
        worksheet = workbook[_sheet_name(table_spec)]
        for item in session.exec(select(table_spec.model)).all():
            worksheet.append(_export_workbook_row_values(session, table_spec, item))

    _add_measure_export_sheet(workbook, session)

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _export_workbook_row_values(session: Session, table_spec: DatasetTableSpec, item: SQLModel) -> list[Any]:
    values = [_serialize_value(getattr(item, column.name, None)) for column in table_spec.columns]
    if table_spec.table_name == "mouse":
        implants = session.exec(
            select(Implant).where(Implant.mouse_id == getattr(item, "id")).order_by(Implant.id)
        ).all()
        for implant in implants[:2]:
            values.extend([
                _serialize_value(implant.id),
                implant.implant_location,
                implant.type,
                _serialize_value(implant.implant_date),
            ])
        missing_implant_slots = 2 - min(len(implants), 2)
        values.extend([None, None, None, None] * missing_implant_slots)
    return values


def _add_measure_export_sheet(workbook: Workbook, session: Session) -> None:
    worksheet = workbook.create_sheet(title="measure")
    columns = _get_model_columns(Measure)
    worksheet.append([column.name for column in columns])
    worksheet.append([_column_note(column) for column in columns])
    worksheet.freeze_panes = "A3"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for item in session.exec(select(Measure)).all():
        worksheet.append([_serialize_value(getattr(item, column.name, None)) for column in columns])


def import_dataset_workbook(
    session: Session,
    workbook_bytes: bytes,
    *,
    filename: str | None = None,
) -> DatasetImportSummary:
    """Import domain data from a workbook containing one sheet per supported table."""
    if not workbook_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty."
        )

    try:
        workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl error types vary by damage
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Excel workbook. Please upload a valid .xlsx file.",
        ) from exc

    summary = _create_import_summary(filename=filename, format_name="xlsx")
    found_supported_sheet = False

    try:
        worksheets = {worksheet.title: worksheet for worksheet in workbook.worksheets}
        for table_spec in get_dataset_table_specs():
            worksheet = worksheets.get(_sheet_name(table_spec))
            if worksheet is None:
                continue

            found_supported_sheet = True
            summary.tables_processed += 1
            header_row = next(
                worksheet.iter_rows(
                    min_row=1, max_row=1, max_col=worksheet.max_column, values_only=True
                ),
                None,
            )
            if header_row is None:
                continue

            expected_columns = _workbook_columns(table_spec)
            expected_headers = tuple(column.name for column in expected_columns)
            actual_headers = _normalize_headers(header_row)
            if actual_headers != expected_headers:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Worksheet {_sheet_name(table_spec)!r} does not match the expected template headers.",
                )

            note_row = next(
                worksheet.iter_rows(
                    min_row=2, max_row=2, max_col=len(expected_headers), values_only=True
                ),
                None,
            )
            has_note_row = note_row is not None and tuple(
                _normalize_text(value) for value in note_row
            ) == tuple(_column_note(column) for column in expected_columns)
            data_start_row = 3 if has_note_row else 2

            for row_number, values in enumerate(
                worksheet.iter_rows(
                    min_row=data_start_row,
                    max_col=len(expected_headers),
                    values_only=True,
                ),
                start=data_start_row,
            ):
                _import_tabular_row(summary, session, table_spec, row_number, values)

    finally:
        workbook.close()

    if not found_supported_sheet:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Workbook does not contain any supported dataset sheets.",
        )

    return summary


def build_dataset_template_csv_zip() -> bytes:
    """Create a ZIP archive with one CSV template per domain table."""
    buffer = BytesIO()
    with ZipFile(buffer, mode="w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("README.txt", _build_readme_text())
        for table_spec in get_dataset_table_specs():
            csv_buffer = StringIO()
            writer = csv.writer(csv_buffer)
            writer.writerow([column.name for column in _workbook_columns(table_spec)])
            archive.writestr(f"{_sheet_name(table_spec)}.csv", csv_buffer.getvalue())

    return buffer.getvalue()


def build_dataset_export_csv_zip(session: Session) -> bytes:
    """Create a ZIP archive with one CSV export per supported domain table."""
    buffer = BytesIO()
    with ZipFile(buffer, mode="w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("README.txt", _build_readme_text())
        for table_spec in get_dataset_table_specs():
            csv_buffer = StringIO()
            writer = csv.writer(csv_buffer)
            writer.writerow([column.name for column in _workbook_columns(table_spec)])
            for item in session.exec(select(table_spec.model)).all():
                writer.writerow(_export_workbook_row_values(session, table_spec, item))
            archive.writestr(f"{_sheet_name(table_spec)}.csv", csv_buffer.getvalue())

        csv_buffer = StringIO()
        writer = csv.writer(csv_buffer)
        measure_columns = _get_model_columns(Measure)
        writer.writerow([column.name for column in measure_columns])
        for item in session.exec(select(Measure)).all():
            writer.writerow([_serialize_value(getattr(item, column.name, None)) for column in measure_columns])
        archive.writestr("measure.csv", csv_buffer.getvalue())

    return buffer.getvalue()


def import_dataset_csv_zip(
    session: Session,
    archive_bytes: bytes,
    *,
    filename: str | None = None,
) -> DatasetImportSummary:
    """Import domain data from a ZIP archive containing one CSV per supported table."""
    if not archive_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty."
        )

    try:
        archive = ZipFile(BytesIO(archive_bytes))
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid ZIP archive. Please upload a valid .zip file.",
        ) from exc

    summary = _create_import_summary(filename=filename, format_name="csv-zip")
    found_supported_file = False

    with archive:
        names = set(archive.namelist())
        for table_spec in get_dataset_table_specs():
            filename_in_archive = f"{_sheet_name(table_spec)}.csv"
            if filename_in_archive not in names:
                continue

            found_supported_file = True
            summary.tables_processed += 1
            with archive.open(filename_in_archive) as handle:
                content = handle.read().decode("utf-8-sig")

            reader = csv.reader(StringIO(content))
            header_row = next(reader, None)
            if header_row is None:
                continue

            expected_columns = _workbook_columns(table_spec)
            expected_headers = tuple(column.name for column in expected_columns)
            if tuple(header_row) != expected_headers:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"CSV file {filename_in_archive!r} does not match the expected template headers.",
                )

            for row_number, values in enumerate(reader, start=2):
                _import_tabular_row(summary, session, table_spec, row_number, values)


    if not found_supported_file:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="ZIP archive does not contain any supported dataset CSV files.",
        )

    return summary


def _get_model_columns(model: type[SQLModel]) -> tuple[DatasetColumnSpec, ...]:
    columns_by_name = {column.name: column for column in model.__table__.columns}
    return tuple(
        DatasetColumnSpec(
            name=field_name,
            required=not columns_by_name[field_name].nullable,
            is_primary_key=columns_by_name[field_name].primary_key,
            foreign_keys=tuple(
                str(foreign_key.column) for foreign_key in columns_by_name[field_name].foreign_keys
            ),
            data_type=_column_data_type(columns_by_name[field_name]),
            format_hint=_column_format_hint(columns_by_name[field_name]),
        )
        for field_name in model.model_fields
        if field_name in columns_by_name
    )


def _sheet_name(table_spec: DatasetTableSpec) -> str:
    return DATASET_SHEET_NAME_ALIASES.get(table_spec.table_name, table_spec.table_name)


def _workbook_columns(table_spec: DatasetTableSpec) -> tuple[DatasetColumnSpec, ...]:
    if table_spec.table_name == "mouse":
        return table_spec.columns + MOUSE_RELATED_COLUMNS
    return table_spec.columns


def _build_base_workbook() -> Workbook:
    workbook = Workbook()
    table_specs = get_dataset_table_specs()
    for index, table_spec in enumerate(table_specs):
        worksheet = (
            workbook.active if index == 0 else workbook.create_sheet(title=_sheet_name(table_spec))
        )
        worksheet.title = _sheet_name(table_spec)
        workbook_columns = _workbook_columns(table_spec)
        header_row = [column.name for column in workbook_columns]
        note_row = [_column_note(column) for column in workbook_columns]
        worksheet.append(header_row)
        worksheet.append(note_row)
        worksheet.freeze_panes = "A3"

        for cell in worksheet[1]:
            cell.font = Font(bold=True)


    return workbook


def _create_import_summary(*, filename: str | None, format_name: str) -> DatasetImportSummary:
    table_counts = {
        table_spec.table_name: EntityImportCounts() for table_spec in get_dataset_table_specs()
    }
    table_counts["implant"] = EntityImportCounts()
    return DatasetImportSummary(
        filename=filename,
        format=format_name,
        table_counts=table_counts,
    )


def _import_tabular_row(
    summary: DatasetImportSummary,
    session: Session,
    table_spec: DatasetTableSpec,
    row_number: int,
    values: tuple[Any, ...] | list[Any],
) -> None:
    if not any(_has_value(value) for value in values):
        summary.rows_skipped += 1
        return

    payload = _build_row_payload(table_spec, values)
    primary_key = _primary_key_value(table_spec, payload)

    try:
        if table_spec.table_name == "mouse":
            action, related_actions = _upsert_mouse_row_with_related(session, table_spec, payload)
        else:
            action = _upsert_row(session, table_spec, payload)
            related_actions = ()
    except (HTTPException, ValidationError, ValueError, SQLAlchemyError) as exc:
        session.rollback()
        summary.rows_failed += 1
        summary.errors.append(
            DatasetImportError(
                table=table_spec.table_name,
                row_number=row_number,
                primary_key=None if primary_key is None else str(primary_key),
                message=_error_message(exc),
            )
        )
        return

    summary.rows_imported += 1
    counts = summary.table_counts[table_spec.table_name]
    if action == "created":
        counts.created += 1
    else:
        counts.updated += 1
    for table_name, related_action in related_actions:
        related_counts = summary.table_counts[table_name]
        if related_action == "created":
            related_counts.created += 1
        else:
            related_counts.updated += 1


def _upsert_mouse_row_with_related(
    session: Session,
    table_spec: DatasetTableSpec,
    payload: dict[str, Any],
) -> tuple[str, tuple[tuple[str, str], ...]]:
    mouse_payload = {
        column.name: payload.get(column.name)
        for column in table_spec.columns
        if column.name in payload
    }
    related_payload = {column.name: payload.get(column.name) for column in MOUSE_RELATED_COLUMNS}

    mouse_action, mouse = _upsert_mouse_row(session, table_spec, mouse_payload)
    related_actions: list[tuple[str, str]] = []

    for implant_number in (1, 2):
        implant_payload = _mouse_implant_payload(related_payload, implant_number)
        if not any(_has_value(value) for value in implant_payload.values()):
            continue

        implant_action, _ = _upsert_mouse_implant(session, mouse.id, implant_payload)
        related_actions.append(("implant", implant_action))

    return mouse_action, tuple(related_actions)


def _upsert_mouse_row(
    session: Session, table_spec: DatasetTableSpec, payload: dict[str, Any]
) -> tuple[str, Mouse]:
    _validate_foreign_keys(session, table_spec, payload)
    primary_key_value = payload.get("id")
    if primary_key_value is not None:
        item = Mouse.model_validate(payload)
        existing = session.get(Mouse, _coerce_primary_key_value(primary_key_value))
        if existing is None:
            return "created", create_item(session, Mouse, item)
        return "updated", update_item(session, Mouse, str(primary_key_value), item)

    pdx_trial_id = payload.get("pdx_trial_id")
    if pdx_trial_id is not None:
        existing = session.exec(select(Mouse).where(Mouse.pdx_trial_id == pdx_trial_id)).first()
        if existing is not None:
            item = Mouse.model_validate({**existing.model_dump(), **payload})
            return "updated", update_item(session, Mouse, str(existing.id), item)

    return "created", create_item(session, Mouse, Mouse.model_validate(payload))


def _mouse_implant_payload(payload: dict[str, Any], implant_number: int) -> dict[str, Any]:
    implant_payload = {
        "id": payload.get(f"implant_{implant_number}_id"),
        "implant_location": payload.get(f"implant_{implant_number}_location"),
        "type": payload.get(f"implant_{implant_number}_type"),
        "implant_date": payload.get(f"implant_{implant_number}_date"),
    }
    return {key: value for key, value in implant_payload.items() if value is not None}


def _upsert_mouse_implant(
    session: Session, mouse_id: UUID, payload: dict[str, Any]
) -> tuple[str, Implant]:
    primary_key_value = payload.get("id")
    item = Implant.model_validate({**payload, "mouse_id": mouse_id})
    if primary_key_value is not None:
        existing = session.get(Implant, _coerce_primary_key_value(primary_key_value))
        if existing is None:
            return "created", create_item(session, Implant, item)
        return "updated", update_item(session, Implant, str(primary_key_value), item)

    existing = session.exec(
        select(Implant)
        .where(Implant.mouse_id == mouse_id)
        .where(Implant.implant_location == payload.get("implant_location"))
        .where(Implant.type == payload.get("type"))
    ).first()
    if existing is None:
        return "created", create_item(session, Implant, item)
    return "updated", update_item(session, Implant, str(existing.id), item)


def _upsert_row(session: Session, table_spec: DatasetTableSpec, payload: dict[str, Any]) -> str:
    model = table_spec.model
    primary_key_column = next(column for column in table_spec.columns if column.is_primary_key)
    primary_key_value = payload.get(primary_key_column.name)
    _validate_foreign_keys(session, table_spec, payload)
    item = model.model_validate(payload)

    if primary_key_value is None:
        create_item(session, model, item)
        return "created"

    existing = session.get(model, _coerce_primary_key_value(primary_key_value))
    if existing is None:
        create_item(session, model, item)
        return "created"

    update_item(session, model, str(primary_key_value), item)
    return "updated"


def _primary_key_value(table_spec: DatasetTableSpec, payload: dict[str, Any]) -> Any:
    primary_key_column = next(column for column in table_spec.columns if column.is_primary_key)
    return payload.get(primary_key_column.name)


def _coerce_primary_key_value(value: Any) -> Any:
    if isinstance(value, str):
        try:
            return UUID(value)
        except ValueError:
            return value
    return value


def _serialize_value(value: Any) -> Any:
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _normalize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        normalized = value.strip()
        if normalized == "":
            return None

        lowered = normalized.casefold()
        if lowered in {"true", "yes", "y", "1", "si", "sí"}:
            return True
        if lowered in {"false", "no", "n", "0"}:
            return False
        return normalized
    return value


def _normalize_headers(values: tuple[Any, ...]) -> tuple[str, ...]:
    normalized = tuple(_normalize_text(value) for value in values)
    return tuple(value for value in normalized if value != "")


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _build_row_payload(
    table_spec: DatasetTableSpec, values: tuple[Any, ...] | list[Any]
) -> dict[str, Any]:
    return _build_payload_from_columns(_workbook_columns(table_spec), values, table_spec=table_spec)


def _build_payload_from_columns(
    columns: tuple[DatasetColumnSpec, ...],
    values: tuple[Any, ...] | list[Any],
    *,
    table_spec: DatasetTableSpec | None = None,
) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for column, value in zip(columns, values, strict=False):
        normalized = _normalize_value(value)
        if normalized is None and column.is_primary_key:
            continue
        if (
            normalized is not None
            and column.data_type == "string"
            and not isinstance(normalized, str)
        ):
            normalized = str(normalized)
        if isinstance(normalized, str) and _should_normalize_passage_identifier(table_spec, column):
            normalized = _normalize_passage_identifier(normalized)
        payload[column.name] = normalized
    return payload


def _should_normalize_passage_identifier(
    table_spec: DatasetTableSpec | None, column: DatasetColumnSpec
) -> bool:
    if table_spec is not None and table_spec.table_name == "passage" and column.name == "id":
        return True
    if column.name in {"passage_id", "parent_passage_id", "pdx_trial_id"}:
        return True
    return "passage.id" in column.foreign_keys


def _normalize_passage_identifier(value: str) -> str:
    return "-".join(part for part in value.strip().split() if part)


def _validate_foreign_keys(
    session: Session, table_spec: DatasetTableSpec, payload: dict[str, Any]
) -> None:
    table_specs_by_name = {spec.table_name: spec for spec in get_dataset_table_specs()}

    for column in table_spec.columns:
        value = payload.get(column.name)
        if value is None:
            continue

        for foreign_key in column.foreign_keys:
            table_name, _, field_name = foreign_key.partition(".")
            referenced_spec = table_specs_by_name.get(table_name)
            if referenced_spec is None:
                continue

            referenced_primary_key = next(
                spec_column for spec_column in referenced_spec.columns if spec_column.is_primary_key
            )
            if field_name != referenced_primary_key.name:
                continue

            referenced_value = _coerce_primary_key_value(value)
            if session.get(referenced_spec.model, referenced_value) is None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Foreign key {column.name} references missing {table_name}.{field_name}: {value}",
                )


def _has_value(value: Any) -> bool:
    return value is not None and (not isinstance(value, str) or value.strip() != "")


def _error_message(exc: Exception) -> str:
    if isinstance(exc, HTTPException):
        return str(exc.detail)
    if isinstance(exc, ValidationError):
        return "; ".join(error["msg"] for error in exc.errors())
    return str(exc)


def _build_readme_text() -> str:
    lines = [
        "TechConnect dataset CSV template",
        "",
        "One CSV file is included per supported domain table.",
        "Keep the header row unchanged.",
        "Preserve primary key values to update existing records; new keys create new records.",
        "Foreign keys must reference an existing or earlier-imported parent row.",
        "Boolean values accept true/false, yes/no, or 1/0. Dates should use YYYY-MM-DD.",
        "Authentication and session tables are excluded from this package.",
        "",
        "Tables:",
    ]
    for table_spec in get_dataset_table_specs():
        lines.append(f"- {table_spec.table_name}: /api/{table_spec.route_prefix}")
    return "\n".join(lines)


def _column_note(column: DatasetColumnSpec) -> str:
    parts: list[str] = []
    if column.is_primary_key:
        parts.append("primary key")
    if column.required:
        parts.append("required")
    else:
        parts.append("optional")
    parts.append(f"type:{column.data_type}")
    if column.format_hint:
        parts.append(f"format:{column.format_hint}")
    parts.extend(f"fk:{foreign_key}" for foreign_key in column.foreign_keys)
    return " | ".join(parts)


def _column_data_type(column) -> str:
    python_type = _column_python_type(column)
    if python_type is UUID:
        return "uuid"
    if python_type is str:
        return "string"
    if python_type is int:
        return "integer"
    if python_type is float:
        return "number"
    if python_type is bool:
        return "boolean"
    if python_type is date:
        return "date"
    if python_type is datetime:
        return "datetime"
    type_name = column.type.__class__.__name__.lower()
    if "string" in type_name:
        return "string"
    return type_name


def _column_format_hint(column) -> str | None:
    python_type = _column_python_type(column)
    if python_type is UUID:
        return "xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx"
    if python_type is bool:
        return "true/false/yes/no/1/0"
    if python_type is date:
        return "YYYY-MM-DD"
    if python_type is datetime:
        return "YYYY-MM-DDTHH:MM:SS"
    return None


def _column_python_type(column) -> type[Any] | None:
    try:
        return column.type.python_type
    except (AttributeError, NotImplementedError):
        return None
