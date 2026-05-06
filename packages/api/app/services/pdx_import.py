"""Import PDX biomodel data from legacy XLSX workbooks."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any

from fastapi import HTTPException, status
from openpyxl import load_workbook
from pydantic import BaseModel, Field, ValidationError
from sqlalchemy.exc import SQLAlchemyError
from sqlmodel import Session, select

from app.services.crud import _check_constraints
from models import Biomodel, Image, Passage, Patient, PDXTrial, Tumor

EXPECTED_TEMPLATE_HEADERS = (
    'BIOMODEL. ID BIOMODELOS',
    'NUEVO TUMOR. Grado',
    'NUEVO TUMOR. Órgano',
    'NUEVO TUMOR. Estadio',
    'NUEVO TUMOR. TNM',
    'NUEVO TUMOR. codigo de tubo',
    'NUEVO TUMOR. Fecha de operación.',
    'PACIENTE. NHC',
    'NUEVO TUMOR. Código de Biobanco',
    'NUEVO TUMOR. Diagnóstico AP',
    'PACIENTE. SEXO',
    'PACIENTE. EDAD',
)
EXPECTED_LEGACY_HEADERS = (
    'TUMOR',
    'BIOBANCO CODE',
    'DATE SURGERY',
    'NHC',
    'CÓDIGO',
    'DIAGNÓSTICO AP BIOPSIA',
)
PASSAGE_BLOCK_WIDTH = 6
PASSAGE_COLUMN_PATTERN = re.compile(r'^px(\d+)$', re.IGNORECASE)
TRUE_MARKERS = {'✅', 'SI', 'SÍ', 'YES', 'TRUE', '1'}
FALSE_MARKERS = {'❌', 'X', '×', 'NO', 'FALSE', '0'}
NON_DATA_MARKERS = TRUE_MARKERS | FALSE_MARKERS


class EntityImportCounts(BaseModel):
    created: int = 0
    updated: int = 0


class ImportRowError(BaseModel):
    sheet: str
    row_number: int
    biomodel_id: str | None = None
    tumor_biobank_code: str | None = None
    patient_nhc: str | None = None
    message: str


class PdxWorkbookImportSummary(BaseModel):
    filename: str | None = None
    sheets_processed: int = 0
    rows_imported: int = 0
    rows_skipped: int = 0
    rows_failed: int = 0
    patients: EntityImportCounts = Field(default_factory=EntityImportCounts)
    tumors: EntityImportCounts = Field(default_factory=EntityImportCounts)
    biomodels: EntityImportCounts = Field(default_factory=EntityImportCounts)
    passages: EntityImportCounts = Field(default_factory=EntityImportCounts)
    pdx_trials: EntityImportCounts = Field(default_factory=EntityImportCounts)
    images: EntityImportCounts = Field(default_factory=EntityImportCounts)
    errors: list[ImportRowError] = Field(default_factory=list)


@dataclass(frozen=True)
class WorkbookLayout:
    name: str
    data_start_row: int


@dataclass
class ParsedPassageBlock:
    passage_number: int
    success: bool | None
    sample_text: str | None
    ffpe: bool | None
    slide_text: str | None
    scanner_magnification: int | None
    image_date: date | None


@dataclass
class ParsedImportRow:
    biomodel_id: str
    tumor_grade: str | None
    tumor_organ: str | None
    tumor_stage: str | None
    tumor_tnm: str | None
    tumor_tube_code: str | None
    operation_date: date | None
    patient_nhc: str
    tumor_biobank_code: str
    tumor_ap_diagnosis: str | None
    patient_sex: str | None
    patient_age: int | None
    incident_note: str | None = None
    passage_blocks: list[ParsedPassageBlock] = field(default_factory=list)


@dataclass
class ImportActions:
    patient: str
    tumor: str
    biomodel: str
    passages: list[str] = field(default_factory=list)
    pdx_trials: list[str] = field(default_factory=list)
    images: list[str] = field(default_factory=list)


def import_pdx_workbook(
    session: Session,
    workbook_bytes: bytes,
    *,
    filename: str | None = None,
) -> PdxWorkbookImportSummary:
    """Import patient, tumor, biomodel, passage, and image records from a workbook."""
    if not workbook_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Uploaded file is empty.')

    try:
        workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl error types vary by file damage
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='Invalid Excel workbook. Please upload a valid .xlsx file.',
        ) from exc

    summary = PdxWorkbookImportSummary(filename=filename)
    found_supported_layout = False

    try:
        for worksheet in workbook.worksheets:
            row1 = next(
                worksheet.iter_rows(min_row=1, max_row=1, max_col=worksheet.max_column, values_only=True),
                None,
            )
            row2 = next(
                worksheet.iter_rows(min_row=2, max_row=2, max_col=worksheet.max_column, values_only=True),
                None,
            )
            if row1 is None or row2 is None:
                continue

            layout = _detect_layout(row1, row2)
            if layout is None:
                continue

            found_supported_layout = True
            summary.sheets_processed += 1
            header_row = tuple(row1)

            for row_number, values in enumerate(
                worksheet.iter_rows(
                    min_row=layout.data_start_row,
                    max_col=worksheet.max_column,
                    values_only=True,
                ),
                start=layout.data_start_row,
            ):
                if not any(_has_value(value) for value in values):
                    summary.rows_skipped += 1
                    continue

                try:
                    parsed_row = _parse_import_row(values, header_row, layout)
                except (ValueError, ValidationError) as exc:
                    summary.rows_failed += 1
                    summary.errors.append(
                        ImportRowError(
                            sheet=worksheet.title,
                            row_number=row_number,
                            message=_error_message(exc),
                        )
                    )
                    continue

                if parsed_row is None:
                    summary.rows_skipped += 1
                    continue

                try:
                    actions = _import_row(session, parsed_row)
                    session.commit()
                except (HTTPException, ValidationError, ValueError, SQLAlchemyError) as exc:
                    session.rollback()
                    summary.rows_failed += 1
                    summary.errors.append(
                        ImportRowError(
                            sheet=worksheet.title,
                            row_number=row_number,
                            biomodel_id=parsed_row.biomodel_id,
                            tumor_biobank_code=parsed_row.tumor_biobank_code,
                            patient_nhc=parsed_row.patient_nhc,
                            message=_error_message(exc),
                        )
                    )
                    continue

                summary.rows_imported += 1
                _increment_action(summary.patients, actions.patient)
                _increment_action(summary.tumors, actions.tumor)
                _increment_action(summary.biomodels, actions.biomodel)
                for action in actions.passages:
                    _increment_action(summary.passages, action)
                for action in actions.pdx_trials:
                    _increment_action(summary.pdx_trials, action)
                for action in actions.images:
                    _increment_action(summary.images, action)
    finally:
        workbook.close()

    if not found_supported_layout:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='Workbook does not contain a supported PDX import layout.',
        )

    return summary


def _detect_layout(row1: tuple[Any, ...], row2: tuple[Any, ...]) -> WorkbookLayout | None:
    normalized_row1 = tuple(_normalize_header(value) for value in row1[: len(EXPECTED_LEGACY_HEADERS)])
    normalized_row2 = tuple(_normalize_header(value) for value in row2[: len(EXPECTED_TEMPLATE_HEADERS)])

    if normalized_row2 == EXPECTED_TEMPLATE_HEADERS:
        return WorkbookLayout(name='template', data_start_row=3)
    if normalized_row1 == EXPECTED_LEGACY_HEADERS:
        return WorkbookLayout(name='legacy', data_start_row=2)
    return None


def _parse_import_row(
    values: tuple[Any, ...],
    row1_headers: tuple[Any, ...],
    layout: WorkbookLayout,
) -> ParsedImportRow | None:
    if layout.name == 'template':
        return _parse_template_row(values, row1_headers)
    if layout.name == 'legacy':
        return _parse_legacy_row(values, row1_headers)
    return None


def _parse_template_row(values: tuple[Any, ...], row1_headers: tuple[Any, ...]) -> ParsedImportRow | None:
    row_values = list(values)
    biomodel_id = _normalize_text(_value_at(row_values, 0))
    patient_nhc = _normalize_text(_value_at(row_values, 7))
    tumor_biobank_code = _normalize_text(_value_at(row_values, 8))

    if not any(_has_value(value) for value in row_values[:12]):
        return None

    missing = [
        label
        for label, value in (
            ('biomodel id', biomodel_id),
            ('patient nhc', patient_nhc),
            ('tumor biobank code', tumor_biobank_code),
        )
        if value is None
    ]
    if missing:
        raise ValueError(f"Missing required values: {', '.join(missing)}")

    return ParsedImportRow(
        biomodel_id=biomodel_id,
        tumor_grade=_normalize_text(_value_at(row_values, 1)),
        tumor_organ=_normalize_text(_value_at(row_values, 2)),
        tumor_stage=_normalize_text(_value_at(row_values, 3)),
        tumor_tnm=_normalize_text(_value_at(row_values, 4)),
        tumor_tube_code=_normalize_text(_value_at(row_values, 5)),
        operation_date=_maybe_date(_value_at(row_values, 6)),
        patient_nhc=patient_nhc,
        tumor_biobank_code=tumor_biobank_code,
        tumor_ap_diagnosis=_normalize_text(_value_at(row_values, 9)),
        patient_sex=_normalize_sex(_value_at(row_values, 10)),
        patient_age=_normalize_int(_value_at(row_values, 11)),
        incident_note=_normalize_text(_value_at(row_values, 15)),
        passage_blocks=_parse_passage_blocks(row_values, row1_headers),
    )


def _parse_legacy_row(values: tuple[Any, ...], row1_headers: tuple[Any, ...]) -> ParsedImportRow | None:
    header_positions = {
        _normalize_header(header): index
        for index, header in enumerate(row1_headers)
        if _normalize_header(header) is not None
    }
    row_values = list(values)

    biomodel_id = _normalize_text(_value_at(row_values, 0))
    patient_nhc = _normalize_text(_header_value(row_values, header_positions, 'NHC'))
    tumor_biobank_code = _normalize_text(_header_value(row_values, header_positions, 'CÓDIGO'))

    if not any(_has_value(value) for value in row_values[: max(8, len(header_positions))]):
        return None

    missing = [
        label
        for label, value in (
            ('biomodel id', biomodel_id),
            ('patient nhc', patient_nhc),
            ('tumor biobank code', tumor_biobank_code),
        )
        if value is None
    ]
    if missing:
        raise ValueError(f"Missing required values: {', '.join(missing)}")

    return ParsedImportRow(
        biomodel_id=biomodel_id,
        tumor_grade=None,
        tumor_organ=None,
        tumor_stage=None,
        tumor_tnm=None,
        tumor_tube_code=_normalize_text(_header_value(row_values, header_positions, 'BIOBANCO CODE')),
        operation_date=_maybe_date(_header_value(row_values, header_positions, 'DATE SURGERY')),
        patient_nhc=patient_nhc,
        tumor_biobank_code=tumor_biobank_code,
        tumor_ap_diagnosis=_normalize_text(_header_value(row_values, header_positions, 'DIAGNÓSTICO AP BIOPSIA')),
        patient_sex=_normalize_sex(_header_value(row_values, header_positions, 'SEXO')),
        patient_age=_normalize_int(_header_value(row_values, header_positions, 'EDAD')),
        incident_note=_normalize_text(_header_value(row_values, header_positions, 'INCIDENCIAS')),
        passage_blocks=_parse_passage_blocks(row_values, row1_headers),
    )


def _parse_passage_blocks(
    row_values: list[Any],
    row1_headers: tuple[Any, ...],
) -> list[ParsedPassageBlock]:
    blocks: list[ParsedPassageBlock] = []
    px_indexes = [index for index, header in enumerate(row1_headers) if _passage_number_from_header(header) is not None]

    for position, index in enumerate(px_indexes):
        header = row1_headers[index]
        passage_number = _passage_number_from_header(header)
        if passage_number is None:
            continue

        next_px_index = px_indexes[position + 1] if position + 1 < len(px_indexes) else len(row1_headers)
        block_headers = row1_headers[index + 1 : next_px_index]
        block_values = row_values[index + 1 : next_px_index]
        if not any(_has_value(value) for value in [_value_at(row_values, index), *block_values]):
            continue

        header_positions = {
            _header_key(block_header): block_index
            for block_index, block_header in enumerate(block_headers)
            if _header_key(block_header) is not None
        }

        success = _normalize_marker(_value_at(row_values, index))
        raw_sample = _block_value(block_values, header_positions, 'SAMPLES')
        raw_ffpe = _block_value(block_values, header_positions, 'FFPE')
        raw_slide = _block_value(block_values, header_positions, 'SLIDES')
        raw_scanner = _block_value(block_values, header_positions, 'SCANNER')

        sample_text = _clean_block_text(raw_sample)
        ffpe = _normalize_marker(raw_ffpe)
        slide_text = _clean_block_text(raw_slide)
        scanner_magnification = _normalize_magnification(raw_scanner)
        image_date = _maybe_date(_block_value(block_values, header_positions, 'BB SAMPLES'))
        if image_date is None:
            image_date = _maybe_date(_block_value(block_values, header_positions, 'BB SAMPLE'))

        if _looks_like_magnification(slide_text) and scanner_magnification is None:
            shifted_ffpe = _normalize_marker(raw_sample)
            shifted_slide = _clean_block_text(raw_ffpe)
            shifted_scanner = _normalize_magnification(raw_slide)
            if shifted_slide is not None and shifted_scanner is not None:
                ffpe = shifted_ffpe if shifted_ffpe is not None else ffpe
                slide_text = shifted_slide
                scanner_magnification = shifted_scanner
                sample_text = None

        if success is None and any(
            value is not None
            for value in (sample_text, ffpe, slide_text, scanner_magnification, image_date)
        ):
            success = True

        blocks.append(
            ParsedPassageBlock(
                passage_number=passage_number,
                success=success,
                sample_text=sample_text,
                ffpe=ffpe,
                slide_text=slide_text,
                scanner_magnification=scanner_magnification,
                image_date=image_date,
            )
        )

    return blocks


def _import_row(session: Session, row: ParsedImportRow) -> ImportActions:
    patient_payload = {'nhc': row.patient_nhc}
    if row.patient_sex is not None:
        patient_payload['sex'] = row.patient_sex
    if row.patient_age is not None:
        patient_payload['age'] = row.patient_age

    tumor_payload: dict[str, Any] = {
        'biobank_code': row.tumor_biobank_code,
        'patient_nhc': row.patient_nhc,
    }
    if row.tumor_tube_code is not None:
        tumor_payload['tube_code'] = row.tumor_tube_code
    if row.tumor_ap_diagnosis is not None:
        tumor_payload['ap_diagnosis'] = row.tumor_ap_diagnosis
    if row.tumor_grade is not None:
        tumor_payload['grade'] = row.tumor_grade
    if row.tumor_organ is not None:
        tumor_payload['organ'] = row.tumor_organ
    if row.tumor_stage is not None:
        tumor_payload['stage'] = row.tumor_stage
    if row.tumor_tnm is not None:
        tumor_payload['tnm'] = row.tumor_tnm
    if row.operation_date is not None:
        tumor_payload['intervention_date'] = row.operation_date

    biomodel_payload: dict[str, Any] = {
        'id': row.biomodel_id,
        'type': 'PDX',
        'tumor_biobank_code': row.tumor_biobank_code,
    }
    if row.operation_date is not None:
        biomodel_payload['creation_date'] = row.operation_date

    actions = ImportActions(
        patient=_upsert_model(session, Patient, row.patient_nhc, patient_payload),
        tumor=_upsert_model(session, Tumor, row.tumor_biobank_code, tumor_payload),
        biomodel=_upsert_model(
            session,
            Biomodel,
            row.biomodel_id,
            biomodel_payload,
            validate_constraints=True,
        ),
    )

    for block in row.passage_blocks:
        passage_id = _passage_id(row.biomodel_id, block.passage_number)
        passage_payload: dict[str, Any] = {
            'id': passage_id,
            'biomodel_id': row.biomodel_id,
        }
        if block.success is not None:
            passage_payload['success'] = block.success
            passage_payload['status'] = block.success

        creation_date = block.image_date or row.operation_date
        if creation_date is not None:
            passage_payload['creation_date'] = creation_date
        if block.image_date is not None:
            passage_payload['biobank_arrival_date'] = block.image_date
            passage_payload['biobank_shipment'] = True

        description_parts = [
            part
            for part in (block.sample_text, row.incident_note if block.passage_number == 1 else None)
            if part
        ]
        if description_parts:
            passage_payload['description'] = ' | '.join(description_parts)

        actions.passages.append(_upsert_model(session, Passage, passage_id, passage_payload))

        if block.ffpe is not None or block.slide_text is not None:
            pdx_trial_payload: dict[str, Any] = {'id': passage_id}
            if block.ffpe is not None:
                pdx_trial_payload['ffpe'] = block.ffpe
            if block.slide_text is not None:
                pdx_trial_payload['he_slide'] = True
            actions.pdx_trials.append(_upsert_model(session, PDXTrial, passage_id, pdx_trial_payload))

        image_payload = {'passage_id': passage_id}
        has_image_data = False
        if block.image_date is not None:
            image_payload['image_date'] = block.image_date
            has_image_data = True
        if block.scanner_magnification is not None:
            image_payload['scanner_magnification'] = block.scanner_magnification
            has_image_data = True
        if block.slide_text is not None:
            image_payload['type'] = block.slide_text
            has_image_data = True
        if has_image_data:
            actions.images.append(_upsert_image(session, passage_id, image_payload))

    return actions


def _upsert_model(
    session: Session,
    model: type[Patient] | type[Tumor] | type[Biomodel] | type[Passage] | type[PDXTrial],
    primary_key: str,
    payload: dict[str, Any],
    *,
    validate_constraints: bool = False,
) -> str:
    existing = session.get(model, primary_key)

    if existing is None:
        if validate_constraints:
            _check_constraints(session, model, payload)
        instance = model.model_validate(payload)
        session.add(instance)
        return 'created'

    merged = {**existing.model_dump(), **payload}
    if validate_constraints:
        _check_constraints(session, model, merged, primary_key)

    validated = model.model_validate(merged)
    existing.sqlmodel_update(validated.model_dump())
    session.add(existing)
    return 'updated'


def _upsert_image(session: Session, passage_id: str, payload: dict[str, Any]) -> str:
    existing = session.exec(select(Image).where(Image.passage_id == passage_id)).first()
    if existing is None:
        session.add(Image.model_validate(payload))
        return 'created'

    validated = Image.model_validate({**existing.model_dump(), **payload})
    existing.sqlmodel_update(validated.model_dump())
    session.add(existing)
    return 'updated'


def _header_value(row_values: list[Any], header_positions: dict[str | None, int], header: str) -> Any:
    index = header_positions.get(header)
    if index is None:
        return None
    return _value_at(row_values, index)


def _block_value(block_values: list[Any], header_positions: dict[str, int], header: str) -> Any:
    index = header_positions.get(header)
    if index is None:
        return None
    return _value_at(block_values, index)


def _value_at(values: list[Any], index: int) -> Any:
    if index < 0 or index >= len(values):
        return None
    return values[index]


def _passage_id(biomodel_id: str, passage_number: int) -> str:
    return f'{biomodel_id}-P{passage_number}'


def _passage_number_from_header(value: Any) -> int | None:
    text = _normalize_text(value)
    if text is None:
        return None
    match = PASSAGE_COLUMN_PATTERN.fullmatch(text)
    if match is None:
        return None
    return int(match.group(1)) + 1


def _clean_block_text(value: Any) -> str | None:
    text = _normalize_text(value)
    if text is None:
        return None
    if text.upper() in NON_DATA_MARKERS:
        return None
    return text


def _header_key(value: Any) -> str | None:
    text = _normalize_text(value)
    if text is None:
        return None
    return text.upper()


def _normalize_header(value: Any) -> str | None:
    return _normalize_text(value)


def _normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
    elif isinstance(value, (int, float)) and float(value).is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()
    return text or None


def _normalize_int(value: Any) -> int | None:
    if value is None or value == '':
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = _normalize_text(value)
    if text is None:
        return None
    return int(float(text))


def _normalize_marker(value: Any) -> bool | None:
    text = _normalize_text(value)
    if text is None:
        return None
    normalized = text.upper()
    if normalized in TRUE_MARKERS:
        return True
    if normalized in FALSE_MARKERS:
        return False
    return None


def _normalize_magnification(value: Any) -> int | None:
    text = _normalize_text(value)
    if text is None:
        return None
    match = re.search(r'(\d+)', text)
    if match is None:
        return None
    return int(match.group(1))


def _looks_like_magnification(value: str | None) -> bool:
    if value is None:
        return False
    return bool(re.fullmatch(r'\d+\s*[xX]', value.strip()))


def _normalize_date(value: Any) -> date | None:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _normalize_text(value)
    if text is None:
        return None

    first_date = re.search(r'\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{2}-\d{2}', text)
    if first_date is None:
        raise ValueError(f'Unsupported date format: {text}')

    token = first_date.group(0)
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(token, fmt).date()
        except ValueError:
            continue

    raise ValueError(f'Unsupported date format: {text}')


def _maybe_date(value: Any) -> date | None:
    text = _normalize_text(value)
    if text is None:
        return None
    if isinstance(value, (datetime, date)):
        return _normalize_date(value)
    if re.search(r'\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{2}-\d{2}', text) is None:
        return None
    return _normalize_date(value)


def _normalize_sex(value: Any) -> str | None:
    text = _normalize_text(value)
    if text is None:
        return None

    normalized = text.upper()
    mapping = {
        'M': 'F',
        'MUJER': 'F',
        'F': 'F',
        'FEMALE': 'F',
        'V': 'M',
        'VARON': 'M',
        'VARÓN': 'M',
        'H': 'M',
        'HOMBRE': 'M',
        'MALE': 'M',
    }
    return mapping.get(normalized, normalized)


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ''
    return True


def _error_message(exc: Exception) -> str:
    if isinstance(exc, HTTPException):
        return str(exc.detail)
    if isinstance(exc, ValidationError):
        return '; '.join(error['msg'] for error in exc.errors())
    return str(exc)


def _increment_action(counts: EntityImportCounts, action: str) -> None:
    setattr(counts, action, getattr(counts, action) + 1)